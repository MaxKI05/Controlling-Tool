import os
import re
from datetime import datetime

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Image as RLImage, SimpleDocTemplate, Spacer, Table, TableStyle

# ──────────────────────────────────────────────────────────────────────────────
# Layout & App-Setup
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Zeitdatenanalyse Dashboard", page_icon="🧠", layout="wide")
APP_VERSION = "v0.0.8"

os.makedirs("history/exports", exist_ok=True)
os.makedirs("history/uploads", exist_ok=True)

# ──────────────────────────────────────────────────────────────────────────────
# Helper-Funktionen
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_data
def load_excel(file):
    return pd.read_excel(file)

def extrahiere_zweck(text: str):
    if isinstance(text, str) and "-" in text:
        zweck_raw = text.split("-")[-1].strip()
        return re.sub(r"^\d+_?", "", zweck_raw)
    return None

def lade_mapping():
    if os.path.exists("mapping.csv"):
        dfm = pd.read_csv("mapping.csv")
        if "Zweck" not in dfm.columns:
            dfm["Zweck"] = ""
        if "Verrechenbarkeit" not in dfm.columns:
            dfm["Verrechenbarkeit"] = "Unbekannt"
        out = dfm[["Zweck", "Verrechenbarkeit"]].copy()
        out["Zweck"] = out["Zweck"].astype(str).str.strip()
        out["Verrechenbarkeit"] = out["Verrechenbarkeit"].fillna("Unbekannt")
        out.drop_duplicates(subset=["Zweck"], inplace=True)
        return out
    return pd.DataFrame(columns=["Zweck", "Verrechenbarkeit"])

def speichere_mapping(df):
    out = df.copy()
    out["Zweck"] = out["Zweck"].astype(str).str.strip()
    out["Verrechenbarkeit"] = out["Verrechenbarkeit"].fillna("Unbekannt")
    out.drop_duplicates(subset=["Zweck"], inplace=True)
    out.to_csv("mapping.csv", index=False)

def lade_kuerzel():
    if os.path.exists("kuerzel.csv"):
        dfk = pd.read_csv("kuerzel.csv")
        if "Name" not in dfk.columns:
            dfk["Name"] = ""
        if "Kürzel" not in dfk.columns:
            dfk["Kürzel"] = ""
        out = dfk[["Name", "Kürzel"]].copy()
        out["Name"] = out["Name"].astype(str).str.strip()
        out["Kürzel"] = out["Kürzel"].astype(str).str.strip()
        out.drop_duplicates(subset=["Name"], inplace=True)
        return out
    return pd.DataFrame(columns=["Name", "Kürzel"])

def speichere_kuerzel(df):
    out = df.copy()
    out["Name"] = out["Name"].astype(str).str.strip()
    out["Kürzel"] = out["Kürzel"].astype(str).str.strip()
    out.dropna(subset=["Name"], inplace=True)
    out.drop_duplicates(subset=["Name"], inplace=True)
    out.to_csv("kuerzel.csv", index=False)

def get_state_df(key, loader):
    if key not in st.session_state or not isinstance(st.session_state[key], pd.DataFrame):
        st.session_state[key] = loader()
    return st.session_state[key]

# ──────────────────────────────────────────────────────────────────────────────
# Session-State initialisieren
# ──────────────────────────────────────────────────────────────────────────────
if "df" not in st.session_state:
    st.session_state["df"] = None
if "mapping_df" not in st.session_state:
    st.session_state["mapping_df"] = lade_mapping()
if "kuerzel_map" not in st.session_state:
    st.session_state["kuerzel_map"] = lade_kuerzel()

# ──────────────────────────────────────────────────────────────────────────────
# Sidebar Navigation
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🧭 Navigation")
    page = st.radio(
        label="Menü",
        options=[
            "🏠 Start",
            "📁 Daten hochladen",
            "🧠 Zweck-Kategorisierung",
            "📊 Analyse & Visualisierung",
            "💰 Abrechnungs-Vergleich",
            "📤 Export",
        ],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown(f"🧠 Max KI Dashboard – {APP_VERSION}")

# ──────────────────────────────────────────────────────────────────────────────
# STARTSEITE
# ──────────────────────────────────────────────────────────────────────────────
if page == "🏠 Start":
    st.title("Willkommen im Zeitdatenanalyse-Dashboard")
    st.markdown(
        """
**Was kann dieses Tool?**

- 📁 Excel-Zeitdaten hochladen
- 🤖 Klassifizierung (Intern/Extern) per Button
- 📊 Interaktive Diagramme
- 📤 Export der Ergebnisse
- 📚 Verlauf vergangener Exporte
"""
    )

    st.markdown("## 📤 Export-Historie")
    export_files = sorted(os.listdir("history/exports"), reverse=True)
    for f in export_files:
        cols = st.columns([8, 1])
        with open(os.path.join("history/exports", f), "rb") as file:
            cols[0].download_button(label=f"⬇️ {f}", data=file.read(), file_name=f)
        if cols[1].button("❌", key=f"del_{f}"):
            os.remove(os.path.join("history/exports", f))
            st.rerun()

# ──────────────────────────────────────────────────────────────────────────────
# DATEN HOCHLADEN – mit automatischem Kürzelimport
# ──────────────────────────────────────────────────────────────────────────────
elif page == "📁 Daten hochladen":
    st.title("📁 Excel-Datei hochladen")
    uploaded_file = st.file_uploader("Lade eine `.xlsx` Datei hoch", type=["xlsx"])

    if uploaded_file:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        save_path = os.path.join("history/uploads", f"upload_{timestamp}.xlsx")
        with open(save_path, "wb") as f:
            f.write(uploaded_file.getvalue())

        df = load_excel(uploaded_file)

        if "Unterprojekt" not in df.columns or "Mitarbeiter" not in df.columns:
            st.error("❌ Spalten 'Unterprojekt' oder 'Mitarbeiter' fehlen.")
        else:
            df["Zweck"] = df["Unterprojekt"].apply(extrahiere_zweck)

            dauer_spalte = None
            for spalte in df.columns:
                if spalte.lower() in ["stunden", "dauer"]:
                    dauer_spalte = spalte
                    break
            if dauer_spalte:
                df["Dauer"] = pd.to_numeric(df[dauer_spalte], errors="coerce").fillna(0)
            else:
                df["Dauer"] = 1.0

            st.session_state["df"] = df

            # ➕ Automatischer Import neuer Mitarbeitenden in kuerzel.csv
            try:
                kuerzel_df = get_state_df("kuerzel_map", lade_kuerzel)
                aktuelle_namen = set(df["Mitarbeiter"].dropna().astype(str).str.strip())
                bekannte_namen = set(kuerzel_df["Name"].astype(str).str.strip())
                neu = sorted(list(aktuelle_namen - bekannte_namen))
                if neu:
                    addon = pd.DataFrame({"Name": neu, "Kürzel": ""})
                    st.session_state["kuerzel_map"] = (
                        pd.concat([kuerzel_df, addon], ignore_index=True)
                          .drop_duplicates(subset=["Name"])
                    )
                    speichere_kuerzel(st.session_state["kuerzel_map"])
                    st.info(f"👥 {len(neu)} neue Mitarbeitende wurden zur Kürzel-Tabelle hinzugefügt.")
            except Exception as e:
                st.warning(f"Konnte neue Mitarbeitende nicht übernehmen: {e}")

            st.success("✅ Datei erfolgreich geladen.")
            st.subheader("📄 Vorschau der Daten")
            st.dataframe(df, use_container_width=True)

    st.markdown("## 📂 Hochgeladene Dateien")
    upload_files = sorted(os.listdir("history/uploads"), reverse=True)
    for f in upload_files:
        with open(os.path.join("history/uploads", f), "rb") as file:
            st.download_button(label=f"📄 {f}", data=file.read(), file_name=f)
            
elif page == "🧠 Zweck-Kategorisierung":
    st.title("🧠 Zweck-Kategorisierung & Mapping")
    df = st.session_state.get("df")
    if df is None or "Zweck" not in df.columns:
        st.warning("⚠️ Bitte zuerst eine Excel-Datei hochladen.")
    else:
        # ---------- 1) Automatisches GPT-Mapping NUR für neue Zwecke ----------
        # Aktuelle & bekannte Zwecke robust ermitteln
        mapping_df = st.session_state.get("mapping_df", lade_mapping())
        aktuelle_zwecke = set(
            df["Zweck"].dropna().astype(str).str.strip()
        )
        bekannte_zwecke = set()
        if not mapping_df.empty and "Zweck" in mapping_df.columns:
            bekannte_zwecke = set(mapping_df["Zweck"].dropna().astype(str).str.strip())

        neue_zwecke = sorted(aktuelle_zwecke - bekannte_zwecke)

        st.markdown(f"🔍 Neue Zwecke im aktuellen Datensatz: **{len(neue_zwecke)}**")

        if neue_zwecke:
            try:
                from utils.gpt import klassifiziere_verrechenbarkeit
                neue_mapping = []
                with st.spinner(f"🧠 {len(neue_zwecke)} neue Zwecke – KI klassifiziert..."):
                    for zweck in neue_zwecke:
                        kat = klassifiziere_verrechenbarkeit(zweck)  # erwartetes Ergebnis: "Intern" oder "Extern"
                        # Falls die KI etwas Unerwartetes zurückgibt, leer lassen -> manuell nachpflegen
                        if kat not in ("Intern", "Extern"):
                            kat = None
                        neue_mapping.append({"Zweck": zweck, "Verrechenbarkeit": kat})

                if neue_mapping:
                    new_df = pd.DataFrame(neue_mapping)
                    mapping_df = pd.concat([mapping_df, new_df], ignore_index=True)
                    mapping_df.drop_duplicates(subset=["Zweck"], keep="last", inplace=True)
                    st.session_state["mapping_df"] = mapping_df
                    speichere_mapping(mapping_df)
                    st.success("✅ Neues Mapping gespeichert.")
            except Exception as e:
                st.warning("⚠️ KI-Mapping konnte nicht automatisch durchgeführt werden. Bitte manuell nachpflegen.")
                # kein st.stop(); wir erlauben manuelle Bearbeitung

        # ---------- 2) Mapping anwenden (merge) ----------
        df = df.drop(columns=["Verrechenbarkeit"], errors="ignore")
        df = df.merge(st.session_state["mapping_df"], on="Zweck", how="left")
        st.session_state["df"] = df

        # ---------- 3) Tabs: aktuelles Mapping & manuell bearbeiten ----------
        tab1, tab2, tab3 = st.tabs(["📋 Aktuelles Mapping", "✍️ Manuell bearbeiten", "👥 Mitarbeiter-Kürzel"])

        with tab1:
            st.caption("Übersicht aller bekannten Zweck→Verrechenbarkeit-Zuordnungen.")
            show_map = st.session_state["mapping_df"].copy()
            if not show_map.empty:
                show_map = show_map.sort_values("Zweck")
            st.dataframe(show_map, use_container_width=True)

        with tab2:
            st.caption("Manuelle Korrektur/Ergänzung des Zweck-Mappings.")
            edited_df = st.data_editor(
                st.session_state["mapping_df"],
                num_rows="dynamic",
                use_container_width=True,
                key="mapping_editor"
            )
            if st.button("💾 Änderungen speichern", key="save_purpose_mapping"):
                # Persistenz + erneutes Anwenden
                edited_df.drop_duplicates(subset=["Zweck"], keep="last", inplace=True)
                st.session_state["mapping_df"] = edited_df
                speichere_mapping(edited_df)

                df_tmp = st.session_state["df"].drop(columns=["Verrechenbarkeit"], errors="ignore")
                df_tmp = df_tmp.merge(edited_df, on="Zweck", how="left")
                st.session_state["df"] = df_tmp

                st.success("✅ Mapping gespeichert & angewendet.")

        # ---------- 4) Mitarbeiter-Kürzel direkt hier pflegen (Session-Only, keine CSV) ----------
        with tab3:
            st.caption("Trage hier manuell die Kürzel zu den Mitarbeitenden ein. Diese werden im Abrechnungs-Vergleich verwendet.")
            # Basis: alle Namen aus den Zeitdaten
            alle_namen = sorted(set(df["Mitarbeiter"].dropna().astype(str)))

            # Bisheriges Mapping aus Session laden (falls vorhanden)
            existing = st.session_state.get("kuerzel_map")
            if existing is None or existing.empty or "Name" not in existing.columns or "Kürzel" not in existing.columns:
                kuerzel_df = pd.DataFrame({"Name": alle_namen, "Kürzel": [""] * len(alle_namen)})
            else:
                # Union aus vorhandenen Namen + neuen Namen, Kürzel wenn vorhanden beibehalten
                kuerzel_df = pd.DataFrame({"Name": alle_namen})
                kuerzel_df = kuerzel_df.merge(existing[["Name", "Kürzel"]], on="Name", how="left").fillna({"Kürzel": ""})

            edited_kuerzel_df = st.data_editor(
                kuerzel_df,
                key="kuerzel_editor",
                use_container_width=True,
                num_rows="dynamic"
            )
            if st.button("💾 Kürzel speichern", key="save_initials"):
                # nur eindeutige Namen, leere Kürzel erlaubt (dann werden diese Personen im Vergleich ignoriert)
                edited_kuerzel_df.drop_duplicates(subset=["Name"], keep="last", inplace=True)
                st.session_state["kuerzel_map"] = edited_kuerzel_df
                st.success("✅ Kürzel wurden gespeichert.")



elif page == "📊 Analyse & Visualisierung":
    st.title("📊 Verrechenbarkeit Gesamtübersicht")

    df = st.session_state.get("df")
    if not isinstance(df, pd.DataFrame):
        st.warning("Bitte zuerst eine Datei hochladen.")
    else:
        if "Verrechenbarkeit" not in df.columns:
            # Mapping nur mergen (kein Auto-Klassifizieren)
            df = df.drop(columns=["Verrechenbarkeit"], errors="ignore")
            df = df.merge(st.session_state["mapping_df"], on="Zweck", how="left")
            st.session_state["df"] = df

        export_df = df[df["Verrechenbarkeit"].isin(["Intern", "Extern"])].copy()
        if export_df.empty:
            st.info("Keine Daten mit 'Intern'/'Extern' vorhanden.")
        else:
            pivot_df = export_df.groupby(["Mitarbeiter", "Verrechenbarkeit"])["Dauer"].sum().unstack(fill_value=0)
            pivot_df["Gesamtstunden"] = pivot_df.sum(axis=1)
            pivot_df["% Intern"] = (pivot_df.get("Intern", 0) / pivot_df["Gesamtstunden"]) * 100
            pivot_df["% Extern"] = (pivot_df.get("Extern", 0) / pivot_df["Gesamtstunden"]) * 100

            export_summary = pivot_df.reset_index()
            export_summary = export_summary[["Mitarbeiter", "Intern", "Extern", "Gesamtstunden", "% Intern", "% Extern"]]
            export_summary[["Intern", "Extern", "Gesamtstunden"]] = export_summary[["Intern", "Extern", "Gesamtstunden"]].round(2)
            export_summary[["% Intern", "% Extern"]] = export_summary[["% Intern", "% Extern"]].round(1)

            st.subheader("📊 Balkendiagramm Intern/Extern pro Mitarbeiter")
            fig, ax = plt.subplots(figsize=(10, 5))
            export_summary.plot(kind="bar", x="Mitarbeiter", y=["Intern", "Extern"], ax=ax)
            ax.set_ylabel("Stunden")
            ax.set_title("Stunden nach Verrechenbarkeit")
            st.pyplot(fig)

            st.subheader("📄 Tabellenansicht")
            st.dataframe(export_summary, use_container_width=True)

elif page == "💰 Abrechnungs-Vergleich":
    st.title("💰 Vergleich: Zeitdaten vs Rechnungsstellung")

    # 1) Parameter
    std_pro_tag = st.number_input(
        "Arbeitsstunden pro Tag (für Umrechnung Stunden → Tage)",
        min_value=1.0, max_value=12.0, value=8.5, step=0.5
    )

    upload = st.file_uploader("Lade eine Abrechnungs-Excel hoch", type=["xlsx"])
    if not upload:
        st.stop()

    # ---------- Excel als Werte laden (Formelergebnisse) ----------
    try:
        import openpyxl
    except Exception:
        st.error("openpyxl fehlt. Bitte `openpyxl` in requirements.txt eintragen.")
        st.stop()

    try:
        wb = openpyxl.load_workbook(upload, data_only=True)
    except Exception as e:
        st.error(f"Abrechnung konnte nicht gelesen werden: {e}")
        st.stop()

    sheetnames = wb.sheetnames
    sheet_name = st.selectbox("Sheet/Monat auswählen", sheetnames, index=len(sheetnames)-1 if sheetnames else 0)
    ws = wb[sheet_name]

    # Hilfen
    import string
    def col_idx_to_letter(idx: int) -> str:
        # 0-basiert → A, B, ..., Z, AA, AB, ...
        idx += 1
        letters = ""
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    def parse_eu_number(x):
        if x is None:
            return None
        s = str(x).strip()
        if not s:
            return None
        s = s.replace("€", "").replace("\u20ac", "").replace("\xa0", "").replace(" ", "")
        if "." in s and "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        try:
            return float(s)
        except:
            return None

    # ---------- komplette Tabelle als DataFrame bauen ----------
    data = []
    max_cols = 0
    for row in ws.iter_rows(values_only=True):
        row_vals = list(row)
        max_cols = max(max_cols, len(row_vals))
        data.append(row_vals)

    # Spaltennamen A, B, C, ... generieren
    cols = [col_idx_to_letter(i) for i in range(max_cols)]
    raw_df = pd.DataFrame(data, columns=cols)

    st.markdown("### 🧱 Tabellenvorschau (ausgewähltes Sheet)")
    st.caption("Tipp: Scrolle nach unten – die Zusammenfassung steht meist im unteren Bereich.")
    st.dataframe(raw_df.tail(200), use_container_width=True)

    # ---------- Baukasten: Header-Handling & Auswahl ----------
    st.markdown("### 🔧 Spalten- & Zeilenauswahl")
    use_header = st.checkbox("Eine Zeile als Kopfzeile verwenden", value=False)
    header_row = None
    df_for_pick = raw_df.copy()

    if use_header:
        header_row = st.number_input("Kopfzeile (1-basiert)", min_value=1, max_value=len(raw_df), value=max(1, len(raw_df)-30))
        # Kopf übernehmen
        header_vals = raw_df.iloc[int(header_row)-1].fillna("").astype(str).str.strip().tolist()
        # leere Header durch Spaltenbuchstaben ersetzen, Duplikate vermeiden
        clean_cols = []
        seen = set()
        for i, h in enumerate(header_vals):
            if not h:
                h = cols[i]
            base = h
            k = 1
            while h in seen:
                k += 1
                h = f"{base}_{k}"
            seen.add(h)
            clean_cols.append(h)
        df_for_pick = raw_df.iloc[int(header_row):].reset_index(drop=True)
        df_for_pick.columns = clean_cols

    # Nutzer wählt Zeilenbereich
    total_rows = len(df_for_pick)
    default_start = max(1, total_rows - 15)
    row_range = st.slider("Zeilenbereich (1-basiert, nach evtl. Kopfzeile)", 1, max(1, total_rows), (default_start, total_rows))
    r_start, r_end = int(row_range[0]), int(row_range[1])
    df_slice = df_for_pick.iloc[r_start-1:r_end].copy()

    # Spaltenauswahl
    kuerzel_col = st.selectbox("Spalte für Kürzel wählen", list(df_slice.columns))
    soll_col    = st.selectbox("Spalte für Einsatztage Abrechnung SOLL wählen", list(df_slice.columns))

    st.markdown("#### 👀 Auswahl-Vorschau")
    preview = df_slice[[kuerzel_col, soll_col]].head(20)
    st.dataframe(preview, use_container_width=True)

    # ---------- Extrahieren & Bereinigen ----------
    abr_records = []
    for _, row in df_slice.iterrows():
        code = str(row.get(kuerzel_col, "")).strip()
        val  = parse_eu_number(row.get(soll_col))
        if code and val is not None:
            abr_records.append({"Kürzel": code, "Einsatztage_SOLL": val})
    abr = pd.DataFrame(abr_records)

    if abr.empty:
        st.error("❌ Keine Einsatztage-Zeilen im gewählten Bereich/Spalten gefunden.")
        st.stop()

    st.success(f"✅ {len(abr)} Zeilen aus dem Bereich extrahiert.")
    st.dataframe(abr, use_container_width=True)

    # ---------- Zeitdaten-IST aus Session (extern) ----------
    df_all = st.session_state.get("df")
    kuerzel_map = st.session_state.get("kuerzel_map", pd.DataFrame())

    if not isinstance(df_all, pd.DataFrame) or df_all.empty:
        st.warning("⚠️ Keine Zeitdaten geladen (Seite '📁 Daten hochladen').")
        st.stop()
    if kuerzel_map.empty or not set(["Name", "Kürzel"]).issubset(kuerzel_map.columns):
        st.warning("⚠️ Kein gültiges Kürzel-Mapping gefunden. Bitte in '🧠 Zweck-Kategorisierung' pflegen.")
        st.stop()

    df_ext = df_all[df_all.get("Verrechenbarkeit").isin(["Extern"])].copy()
    if df_ext.empty:
        st.info("Keine externen Zeitdaten vorhanden.")
        st.stop()

    df_ext_group = (
        df_ext.groupby("Mitarbeiter", as_index=False)["Dauer"]
              .sum()
              .rename(columns={"Dauer": "Externe_Stunden"})
    )
    df_ext_map = (
        df_ext_group.merge(kuerzel_map, left_on="Mitarbeiter", right_on="Name", how="left")
                    .dropna(subset=["Kürzel"])
    )
    df_ext_map["Kürzel"] = df_ext_map["Kürzel"].astype(str).str.strip()
    ist_by_k = df_ext_map.groupby("Kürzel", as_index=False)["Externe_Stunden"].sum()
    ist_by_k["Tage_IST"] = ist_by_k["Externe_Stunden"] / float(std_pro_tag)

    # ---------- Mergen & Ausgabe ----------
    merged = abr.groupby("Kürzel", as_index=False).sum(numeric_only=True).merge(
        ist_by_k, on="Kürzel", how="outer"
    ).fillna(0)
    merged["Diff_Tage"] = merged["Tage_IST"] - merged["Einsatztage_SOLL"]

    out = merged.copy()
    if "Externe_Stunden" in out:
        out["Externe_Stunden"] = out["Externe_Stunden"].round(2)
    out["Tage_IST"] = out["Tage_IST"].round(2)
    out["Einsatztage_SOLL"] = out["Einsatztage_SOLL"].round(2)
    out["Diff_Tage"] = out["Diff_Tage"].round(2)

    st.subheader("📊 Vergleichstabelle")
    st.dataframe(
        out[["Kürzel", "Externe_Stunden", "Tage_IST", "Einsatztage_SOLL", "Diff_Tage"]]
          .sort_values("Kürzel"),
        use_container_width=True
    )

    st.caption(
        f"Quelle: Sheet **{sheet_name}**, Zeilen {r_start}–{r_end}, Spalten **{kuerzel_col}** (Kürzel) und **{soll_col}** (Einsatztage_SOLL). "
        f"Zeitdaten extern → Externe_Stunden ÷ {std_pro_tag:g} = Tage_IST. Diff_Tage = Tage_IST − Einsatztage_SOLL."
    )


elif page == "📤 Export":
    st.title("📤 Datenexport")

    df = st.session_state.get("df")
    if not isinstance(df, pd.DataFrame):
        st.info("Bitte zuerst Daten hochladen und klassifizieren.")
    else:
        export_df = df.copy()

        if "Verrechenbarkeit" not in export_df.columns:
            export_df = export_df.merge(st.session_state["mapping_df"], on="Zweck", how="left")

        if "Dauer" not in export_df.columns:
            st.error("❌ Keine 'Dauer'-Spalte gefunden.")
            st.stop()

        export_df = export_df[export_df["Verrechenbarkeit"].isin(["Intern", "Extern"])]

        pivot_df = export_df.groupby(["Mitarbeiter", "Verrechenbarkeit"])["Dauer"].sum().unstack(fill_value=0)
        pivot_df["Gesamtstunden"] = pivot_df.sum(axis=1)
        pivot_df["% Intern"] = (pivot_df.get("Intern", 0) / pivot_df["Gesamtstunden"]) * 100
        pivot_df["% Extern"] = (pivot_df.get("Extern", 0) / pivot_df["Gesamtstunden"]) * 100

        export_summary = pivot_df.reset_index()
        export_summary = export_summary[["Mitarbeiter", "Intern", "Extern", "Gesamtstunden", "% Intern", "% Extern"]]
        export_summary[["Intern", "Extern", "Gesamtstunden"]] = export_summary[["Intern", "Extern", "Gesamtstunden"]].round(2)
        export_summary[["% Intern", "% Extern"]] = export_summary[["% Intern", "% Extern"]].round(1)

        # Diagramm für PDF speichern
        fig, ax = plt.subplots(figsize=(10, 5))
        export_summary.plot(kind="bar", x="Mitarbeiter", y=["Intern", "Extern"], ax=ax)
        ax.set_ylabel("Stunden")
        ax.set_title("Stunden nach Verrechenbarkeit")
        fig.tight_layout()

        image_path = "temp_chart.png"
        fig.savefig(image_path)

        # PDF bauen
        pdf_path = f"history/exports/bericht_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        elements = [RLImage(image_path, width=500, height=300), Spacer(1, 12)]

        table_data = [export_summary.columns.tolist()] + export_summary.values.tolist()
        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)

        with open(pdf_path, "rb") as f:
            st.download_button(
                "⬇️ PDF-Bericht herunterladen",
                data=f.read(),
                file_name=os.path.basename(pdf_path),
                mime="application/pdf",
            )

